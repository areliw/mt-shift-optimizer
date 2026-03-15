# main.py —- FastAPI entry for MT Shift Optimizer

import collections
import io
import csv
import logging
import os
import re as _re
import sqlite3
import time
from pathlib import Path

from fastapi import Body, Depends, FastAPI, Header, HTTPException, Query, APIRouter, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, field_validator

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
)
logger = logging.getLogger("mt_shift_optimizer")

from database import (
    init_master_db,
    create_workspace,
    delete_workspace,
    get_workspace,
    list_workspaces,
    set_workspace_context,
    verify_workspace_token,
    get_mt_list,
    get_shift_list,
    get_num_days,
    set_num_days,
    get_schedule_start_date,
    set_schedule_start_date,
    get_holiday_dates,
    set_holiday_dates,
    get_latest_schedule,
    get_schedule,
    save_schedule,
    update_slot_staff,
    list_staff,
    get_staff,
    list_shifts,
    create_staff,
    update_staff,
    delete_staff,
    list_skill_catalog,
    add_skill_catalog,
    remove_skill_catalog,
    rename_skill_catalog,
    get_skill_levels,
    set_skill_levels,
    list_title_catalog,
    add_title_catalog,
    remove_title_catalog,
    list_time_window_catalog,
    add_time_window_catalog,
    remove_time_window_catalog,
    create_shift,
    update_shift,
    delete_shift,
    move_shift,
    create_shift_from_template,
    apply_template,
    clear_all,
    list_staff_pairs,
    add_staff_pair,
    remove_staff_pair,
    export_all_data,
    import_all_data,
    swap_slots,
    increment_schedule_run_count,
    get_schedule_run_count,
)
from scheduler import generate_schedule, diagnose_infeasible, DUMMY_WORKER
from ortools.sat.python import cp_model
from datetime import datetime, timedelta


def _check_staff_off_day_warnings(staff_name: str, day: int) -> list[str]:
    """เช็คว่าคนนี้หยุดวันนี้ไหม คืน list ของ warning (ว่าง = ไม่มีปัญหา)"""
    warnings = []
    # หา staff จาก mt_list by name
    mt_list = get_mt_list()
    mt = next((m for m in mt_list if m["name"] == staff_name), None)
    if not mt:
        return warnings
    start_str = get_schedule_start_date()
    if not start_str:
        return warnings
    try:
        start_date = datetime.strptime(start_str.strip()[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return warnings
    cal_date = start_date + timedelta(days=day)
    # เช็ค off_days (weekday)
    off_weekdays = set(mt.get("off_days") or [])
    if cal_date.weekday() in off_weekdays:
        day_names = ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"]
        warnings.append(f"'{staff_name}' ตั้งหยุดทุกวัน{day_names[cal_date.weekday()]} แต่วันที่ {cal_date.isoformat()} เป็นวัน{day_names[cal_date.weekday()]}")
    # เช็ค off_days_of_month
    off_month = mt.get("off_days_of_month") or []
    if cal_date.day in off_month:
        warnings.append(f"'{staff_name}' ตั้งหยุดวันที่ {cal_date.day} ของเดือน แต่ถูกจัดลงวันที่ {cal_date.isoformat()}")
    return warnings

# Initialize master DB (workspace registry) + migrate old data
init_master_db()

app = FastAPI(title="MT Shift Optimizer")

# ---------------------------------------------------------------------------
# CORS — configure allowed origins via ALLOWED_ORIGINS env var
# (comma-separated, e.g. "https://app.example.com,https://admin.example.com")
# Defaults to "*" in dev; set a restrictive list in production.
# ---------------------------------------------------------------------------
_raw_origins = os.environ.get("ALLOWED_ORIGINS", "").strip()
_ALLOWED_ORIGINS = [o.strip() for o in _raw_origins.split(",") if o.strip()] or ["*"]
if "*" in _ALLOWED_ORIGINS:
    logger.warning("CORS is open to all origins. Set ALLOWED_ORIGINS env var in production.")
app.add_middleware(
    CORSMiddleware,
    allow_origins=_ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Rate limiting — simple in-memory sliding-window per client IP
# ---------------------------------------------------------------------------
_RATE_LIMIT_WINDOW = int(os.environ.get("RATE_LIMIT_WINDOW", "60"))   # seconds
_RATE_LIMIT_MAX = int(os.environ.get("RATE_LIMIT_MAX", "120"))         # requests / window
_RATE_TRACKED_IP_MAX = int(os.environ.get("RATE_TRACKED_IP_MAX", "10000"))
_RATE_COUNTER_GC_INTERVAL = int(os.environ.get("RATE_COUNTER_GC_INTERVAL", "30"))
_rate_counters: dict[str, collections.deque[float]] = {}
_rate_last_seen: dict[str, float] = {}
_rate_last_gc_at = 0.0


def _rate_limit_gc(now: float, window_start: float):
    """Prune stale per-IP buckets and cap total tracked IPs."""
    stale_ips = []
    for ip, dq in list(_rate_counters.items()):
        while dq and dq[0] < window_start:
            dq.popleft()
        last_seen = _rate_last_seen.get(ip, 0.0)
        if not dq and last_seen < window_start:
            stale_ips.append(ip)

    for ip in stale_ips:
        _rate_counters.pop(ip, None)
        _rate_last_seen.pop(ip, None)

    if len(_rate_counters) > _RATE_TRACKED_IP_MAX:
        # Evict least-recently-seen IPs first to keep memory bounded.
        overflow = len(_rate_counters) - _RATE_TRACKED_IP_MAX
        for ip, _ in sorted(_rate_last_seen.items(), key=lambda kv: kv[1])[:overflow]:
            _rate_counters.pop(ip, None)
            _rate_last_seen.pop(ip, None)


@app.middleware("http")
async def rate_limit_middleware(request: Request, call_next):
    global _rate_last_gc_at
    client_ip = (request.client.host if request.client else "unknown")
    now = time.monotonic()
    window_start = now - _RATE_LIMIT_WINDOW
    if now - _rate_last_gc_at >= _RATE_COUNTER_GC_INTERVAL:
        _rate_limit_gc(now, window_start)
        _rate_last_gc_at = now

    timestamps = _rate_counters.get(client_ip)
    if timestamps is None:
        timestamps = collections.deque(maxlen=max(_RATE_LIMIT_MAX + 1, 2))
        _rate_counters[client_ip] = timestamps

    while timestamps and timestamps[0] < window_start:
        timestamps.popleft()

    _rate_last_seen[client_ip] = now
    if len(timestamps) >= _RATE_LIMIT_MAX:
        logger.warning("Rate limit exceeded for %s", client_ip)
        return JSONResponse(
            status_code=429,
            content={"detail": "Rate limit exceeded. Please slow down."},
        )
    timestamps.append(now)
    return await call_next(request)


# ---------------------------------------------------------------------------
# API Key authentication
# Set API_KEY env var to enable.  All /api/* and /w/* routes require the key.
# If API_KEY is unset the server runs in open dev mode (warned at startup).
# Clients must send:  X-API-Key: <key>
# ---------------------------------------------------------------------------
_API_KEY = os.environ.get("API_KEY", "").strip()
if not _API_KEY:
    logger.warning(
        "API_KEY env var is not set — all endpoints are publicly accessible. "
        "Set API_KEY before deploying to production."
    )


async def verify_api_key(x_api_key: str | None = Header(default=None)):
    """FastAPI dependency: validate X-API-Key header when API_KEY is configured."""
    if not _API_KEY:
        return  # dev mode — skip auth
    if not x_api_key or x_api_key != _API_KEY:
        logger.warning("Rejected request: invalid or missing X-API-Key")
        raise HTTPException(status_code=401, detail="Invalid or missing API key")


# ---------------------------------------------------------------------------
# Input validation helpers
# ---------------------------------------------------------------------------
_HTML_CHARS_RE = _re.compile(r"[<>]")
_MAX_NAME_LEN = 100


def _validate_display_name(value: str, field: str = "name") -> str:
    """Strip whitespace, reject HTML chars, enforce length limit."""
    v = (value or "").strip()
    if not v:
        raise ValueError(f"{field} ต้องไม่ว่าง")
    if len(v) > _MAX_NAME_LEN:
        raise ValueError(f"{field} ยาวเกินไป (สูงสุด {_MAX_NAME_LEN} ตัวอักษร)")
    if _HTML_CHARS_RE.search(v):
        raise ValueError(f"{field} ต้องไม่มีอักขระ < หรือ >")
    return v

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# --- Workspace dependency (async so contextvar propagates to sync endpoint threads) ---
# โหมด dev: ไม่ตั้ง API_KEY = ข้าม workspace token (ใช้ได้เลย)
# โหมด production: ตั้ง API_KEY = ต้องมี workspace token
_ws_auth_env = os.environ.get("DISABLE_WORKSPACE_AUTH", "").lower()
_DISABLE_WORKSPACE_AUTH = (
    _ws_auth_env in ("1", "true", "yes")
    or (not _API_KEY and _ws_auth_env != "0" and _ws_auth_env != "false")
)
if _DISABLE_WORKSPACE_AUTH:
    logger.info("Workspace token check OFF (dev mode). Set API_KEY + DISABLE_WORKSPACE_AUTH=0 for production.")


async def workspace_dep(workspace_id: str, x_workspace_token: str | None = Header(default=None)):
    """FastAPI dependency: validate workspace exists AND token matches"""
    ws = get_workspace(workspace_id)
    if ws is None:
        raise HTTPException(status_code=404, detail="Workspace not found")
    if not _DISABLE_WORKSPACE_AUTH and not verify_workspace_token(workspace_id, x_workspace_token or ""):
        logger.warning("Rejected workspace access: id=%s invalid token", workspace_id)
        raise HTTPException(status_code=403, detail="Invalid or missing workspace token")
    set_workspace_context(workspace_id)
    return ws


# --- Pydantic models ---
class StaffCreate(BaseModel):
    name: str
    title: str = ""
    off_days: list[int] = []
    off_days_of_month: list[int] = []
    skills: list[str] = []
    time_windows: list[str] = []
    min_shifts_per_month: int | None = None
    max_shifts_per_month: int | None = None
    min_gap_days: int | None = None
    min_gap_shifts: list[str] = []
    min_gap_rules: list[dict] = []
    shift_day_rules: list[dict] = []
    shift_limits: dict[str, dict[str, int | None]] = {}

    @field_validator("name")
    @classmethod
    def validate_name(cls, v: str) -> str:
        return _validate_display_name(v, "ชื่อบุคลากร")

    @field_validator("off_days", mode="before")
    @classmethod
    def validate_off_days(cls, v: list) -> list:
        return [int(d) for d in v if 0 <= int(d) <= 6]

    @field_validator("off_days_of_month", mode="before")
    @classmethod
    def validate_off_days_of_month(cls, v: list) -> list:
        return [int(d) for d in v if 1 <= int(d) <= 31]


class StaffUpdate(BaseModel):
    name: str
    title: str = ""
    off_days: list[int] = []
    off_days_of_month: list[int] = []
    skills: list[str] = []
    time_windows: list[str] = []
    skill_levels: dict[str, int] = {}
    min_shifts_per_month: int | None = None
    max_shifts_per_month: int | None = None
    min_gap_days: int | None = None
    min_gap_shifts: list[str] = []
    min_gap_rules: list[dict] = []
    shift_day_rules: list[dict] = []
    shift_limits: dict[str, dict[str, int | None]] = {}

    @field_validator("name")
    @classmethod
    def validate_name(cls, v: str) -> str:
        return _validate_display_name(v, "ชื่อบุคลากร")

    @field_validator("off_days", mode="before")
    @classmethod
    def validate_off_days(cls, v: list) -> list:
        return [int(d) for d in v if 0 <= int(d) <= 6]

    @field_validator("off_days_of_month", mode="before")
    @classmethod
    def validate_off_days_of_month(cls, v: list) -> list:
        return [int(d) for d in v if 1 <= int(d) <= 31]


class PositionItem(BaseModel):
    name: str
    constraint_note: str | None = None
    regular_only: bool | None = None
    slot_count: int = 1  # จำนวนคน (เช่น ช่อง 1-10 = 10 คน)
    time_window_name: str | None = None  # ช่วงเวลาที่ช่องต้องการ เช่น 06:30-12:00
    required_skill: str | None = None  # ทักษะที่ต้องการ เช่น "เจาะเลือด"
    min_skill_level: int = 0  # ระดับ skill ขั้นต่ำ (0=ไม่กำหนด, 1=ต่ำ, 2=กลาง, 3=สูง)
    allowed_titles: list[str] = []  # ฉายาที่อนุญาต (ว่าง = ทุกฉายา)
    max_per_week: int = 0  # สูงสุดกี่ครั้ง/สัปดาห์ (0 = ไม่จำกัด)
    active_weekdays: str | None = None  # เปิดเฉพาะวัน (0=จ … 6=อา) เช่น "6" = อาทิตย์เท่านั้น ว่าง = ทุกวันที่กะเปิด
    holiday_mode: str = "all"  # all | non_holiday_only | holiday_only

    @field_validator("name")
    @classmethod
    def validate_name(cls, v: str) -> str:
        return _validate_display_name(v, "ชื่อตำแหน่ง")

    @field_validator("holiday_mode")
    @classmethod
    def validate_holiday_mode(cls, v: str) -> str:
        mode = (v or "all").strip() or "all"
        if mode not in ("all", "non_holiday_only", "holiday_only"):
            return "all"
        return mode


class TimeWindowCreate(BaseModel):
    name: str  # เช่น 06:30-12:00 (หรือให้ระบบสร้างจาก start_time-end_time)
    start_time: str = ""  # HH:MM
    end_time: str = ""    # HH:MM

    @field_validator("name")
    @classmethod
    def validate_name(cls, v: str) -> str:
        return _validate_display_name(v, "ชื่อช่วงเวลา")


class SkillCreate(BaseModel):
    name: str

    @field_validator("name")
    @classmethod
    def validate_name(cls, v: str) -> str:
        return _validate_display_name(v, "ชื่อทักษะ")


class TitleCreate(BaseModel):
    name: str
    type: str = "fulltime"  # fulltime | parttime

    @field_validator("name")
    @classmethod
    def validate_name(cls, v: str) -> str:
        return _validate_display_name(v, "ชื่อฉายา")


class ShiftCreate(BaseModel):
    name: str
    donor: int = 0
    xmatch: int = 0
    positions: list[PositionItem] | None = None
    active_days: str | None = None
    active_days_of_month: list[int] = []
    include_holidays: bool = False
    title_requirements: list[dict] = []

    @field_validator("name")
    @classmethod
    def validate_name(cls, v: str) -> str:
        return _validate_display_name(v, "ชื่อกะ")

    @field_validator("active_days_of_month", mode="before")
    @classmethod
    def validate_active_days_of_month(cls, v: list) -> list:
        return [int(d) for d in v if 1 <= int(d) <= 31]


class ShiftUpdate(BaseModel):
    name: str
    donor: int = 0
    xmatch: int = 0
    positions: list[PositionItem] | None = None
    active_days: str | None = None
    active_days_of_month: list[int] = []
    include_holidays: bool = False
    title_requirements: list[dict] = []

    @field_validator("name")
    @classmethod
    def validate_name(cls, v: str) -> str:
        return _validate_display_name(v, "ชื่อกะ")

    @field_validator("active_days_of_month", mode="before")
    @classmethod
    def validate_active_days_of_month(cls, v: list) -> list:
        return [int(d) for d in v if 1 <= int(d) <= 31]


class ShiftMoveBody(BaseModel):
    direction: str  # "up" | "down"


class NumDaysUpdate(BaseModel):
    value: int


class SettingsUpdate(BaseModel):
    num_days: int | None = None
    schedule_start_date: str | None = None  # YYYY-MM-DD or "" to clear
    holiday_dates: str | None = None  # comma-separated YYYY-MM-DD


class StaffPairCreate(BaseModel):
    staff_id_1: int
    staff_id_2: int
    pair_type: str  # "together", "apart", or "depends_on"
    shift_names: list[str] | None = None  # ว่าง = ทุกกะ, มีค่า = เฉพาะกะนั้น


class StaffPairBatchItem(BaseModel):
    name_1: str
    name_2: str
    pair_type: str = "together"
    shift_names: list[str] | None = None


class ScheduleRunBody(BaseModel):
    """Optional: use form values for this run and sync to DB."""
    num_days: int | None = None
    schedule_start_date: str | None = None


class SlotAssign(BaseModel):
    """Manual assignment: แทนที่ dummy slot ด้วย staff จริง"""
    day: int
    shift_name: str
    position: str
    slot_index: int = 0
    staff_name: str


class SlotSwap(BaseModel):
    """Swap staff between two slots"""
    day_a: int
    shift_name_a: str
    position_a: str
    slot_index_a: int = 0
    day_b: int
    shift_name_b: str
    position_b: str
    slot_index_b: int = 0


class WorkspaceCreate(BaseModel):
    name: str = ""

    @field_validator("name", mode="before")
    @classmethod
    def validate_name(cls, v: str) -> str:
        v = (v or "").strip()
        if v and len(v) > _MAX_NAME_LEN:
            raise ValueError(f"ชื่อ workspace ยาวเกินไป (สูงสุด {_MAX_NAME_LEN} ตัวอักษร)")
        if _HTML_CHARS_RE.search(v):
            raise ValueError("ชื่อต้องไม่มีอักขระ < หรือ >")
        return v


# ==========================================
# Workspace-scoped router: /w/{workspace_id}/api/...
# ==========================================
ws_router = APIRouter(dependencies=[Depends(workspace_dep), Depends(verify_api_key)])


# --- API: Staff ---
@ws_router.get("/api/staff")
def api_list_staff():
    return list_staff()


@ws_router.get("/api/staff/{staff_id:int}")
def api_get_staff(staff_id: int):
    staff = get_staff(staff_id)
    if staff is None:
        raise HTTPException(status_code=404, detail="Staff not found.")
    return staff


@ws_router.get("/api/time-windows")
def api_list_time_windows():
    return list_time_window_catalog()


@ws_router.post("/api/time-windows")
def api_add_time_window(body: TimeWindowCreate):
    name = (body.name or "").strip()
    start_time = (body.start_time or "").strip()
    end_time = (body.end_time or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="ชื่อช่วงเวลาต้องไม่ว่าง")
    if not start_time or not end_time:
        if "-" in name:
            parts = name.split("-", 1)
            start_time = start_time or (parts[0].strip() if parts else "")
            end_time = end_time or (parts[1].strip() if len(parts) > 1 else "")
        if not start_time or not end_time:
            raise HTTPException(status_code=400, detail="ระบุ start_time และ end_time (เช่น 06:30, 12:00) หรือใช้รูปแบบ 06:30-12:00 ใน name")
    add_time_window_catalog(name, start_time, end_time)
    return {"name": name, "start_time": start_time, "end_time": end_time}


@ws_router.delete("/api/time-windows/{name:path}")
def api_remove_time_window(name: str):
    remove_time_window_catalog(name)
    return {"ok": True}


@ws_router.post("/api/staff")
def api_create_staff(body: StaffCreate):
    try:
        sid = create_staff(
            body.name,
            body.off_days,
            body.skills,
            body.title,
            body.off_days_of_month,
            body.time_windows,
            min_shifts_per_month=body.min_shifts_per_month,
            max_shifts_per_month=body.max_shifts_per_month,
            min_gap_days=body.min_gap_days,
            min_gap_shifts=body.min_gap_shifts,
            min_gap_rules=body.min_gap_rules,
            shift_day_rules=body.shift_day_rules,
            shift_limits=body.shift_limits,
        )
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail=f"ชื่อ '{body.name}' มีอยู่แล้ว กรุณาใช้ชื่ออื่น")
    from database import get_title_type
    stype = get_title_type(body.title or "")
    return {"id": sid, "name": body.name, "type": stype, "title": body.title or "", "off_days": body.off_days, "skills": body.skills}


@ws_router.put("/api/staff/{staff_id:int}")
def api_update_staff(staff_id: int, body: StaffUpdate):
    try:
        update_staff(
            staff_id,
            body.name,
            body.off_days,
            body.skills,
            body.title,
            body.off_days_of_month,
            body.time_windows,
            skill_levels=body.skill_levels,
            min_shifts_per_month=body.min_shifts_per_month,
            max_shifts_per_month=body.max_shifts_per_month,
            min_gap_days=body.min_gap_days,
            min_gap_shifts=body.min_gap_shifts,
            min_gap_rules=body.min_gap_rules,
            shift_day_rules=body.shift_day_rules,
            shift_limits=body.shift_limits,
        )
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail=f"ชื่อ '{body.name}' มีอยู่แล้ว กรุณาใช้ชื่ออื่น")
    from database import get_title_type
    stype = get_title_type(body.title or "")
    return {"id": staff_id, "name": body.name, "type": stype, "title": body.title or "", "off_days": body.off_days, "skills": body.skills}


@ws_router.delete("/api/staff/{staff_id:int}")
def api_delete_staff(staff_id: int):
    delete_staff(staff_id)
    return {"ok": True}


# --- API: Skills (รายการทักษะสำหรับใส่ให้บุคลากร) ---
@ws_router.get("/api/skills")
def api_list_skills():
    return list_skill_catalog()


@ws_router.post("/api/skills")
def api_add_skill(body: SkillCreate):
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="ชื่อทักษะต้องไม่ว่าง")
    add_skill_catalog(name)
    return {"name": name}


@ws_router.put("/api/skills/{old_name}")
def api_rename_skill(old_name: str, body: SkillCreate):
    old = (old_name or "").strip()
    new = (body.name or "").strip()
    if not new:
        raise HTTPException(status_code=400, detail="ชื่อทักษะต้องไม่ว่าง")
    if not old or old == new:
        return {"name": new}
    try:
        rename_skill_catalog(old, new)
    except Exception as exc:  # sqlite3.IntegrityError หรืออื่นๆ
        raise HTTPException(status_code=400, detail="ไม่สามารถเปลี่ยนชื่อทักษะได้ (อาจมีชื่อซ้ำ)") from exc
    return {"name": new}


@ws_router.delete("/api/skills/{name}")
def api_remove_skill(name: str):
    remove_skill_catalog(name)
    return {"ok": True}


class SkillLevelsUpdate(BaseModel):
    levels: list[str]


@ws_router.get("/api/skills/{name}/levels")
def api_get_skill_levels(name: str):
    return get_skill_levels(name)


@ws_router.put("/api/skills/{name}/levels")
def api_set_skill_levels(name: str, body: SkillLevelsUpdate):
    labels = [l.strip() for l in body.levels if l.strip()]
    if not labels:
        raise HTTPException(status_code=400, detail="ต้องมีอย่างน้อย 1 ระดับ")
    set_skill_levels(name, labels)
    return get_skill_levels(name)


# --- API: Titles (ฉายา/ตำแหน่ง รวมประเภท) ---
@ws_router.get("/api/titles")
def api_list_titles():
    return list_title_catalog()


@ws_router.post("/api/titles")
def api_add_title(body: TitleCreate):
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="ชื่อฉายาต้องไม่ว่าง")
    stype = (body.type or "fulltime").lower()
    if stype not in ("fulltime", "parttime"):
        stype = "fulltime"
    try:
        add_title_catalog(name, stype)
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="ชื่อฉายาซ้ำ มีอยู่แล้ว กรุณาใช้ชื่ออื่น")
    return {"name": name, "type": stype}


@ws_router.delete("/api/titles/{name:path}")
def api_remove_title(name: str):
    remove_title_catalog(name)
    return {"ok": True}


# --- API: Shifts ---
@ws_router.get("/api/shifts")
def api_list_shifts():
    return list_shifts()


@ws_router.post("/api/shifts")
def api_create_shift(body: ShiftCreate):
    positions = None
    if body.positions is not None:
        positions = [{"name": p.name, "constraint_note": p.constraint_note or "", "regular_only": p.regular_only or False, "slot_count": max(1, p.slot_count or 1), "time_window_name": (p.time_window_name or "").strip() or None, "required_skill": (p.required_skill or "").strip() or None, "min_skill_level": max(0, int(p.min_skill_level or 0)), "allowed_titles": list(p.allowed_titles or []), "max_per_week": max(0, int(p.max_per_week or 0)), "active_weekdays": (p.active_weekdays or "").strip() or None, "holiday_mode": (p.holiday_mode or "all").strip() or "all"} for p in body.positions]
    try:
        sid = create_shift(body.name, body.donor, body.xmatch, positions=positions, active_days=body.active_days, active_days_of_month=body.active_days_of_month, include_holidays=body.include_holidays, title_requirements=body.title_requirements)
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="ชื่อกะซ้ำ มีอยู่แล้ว กรุณาใช้ชื่ออื่น")
    out = {"id": sid, "name": body.name}
    if positions is not None:
        out["positions"] = [{"name": p["name"], "constraint_note": p["constraint_note"], "regular_only": p["regular_only"], "slot_count": p.get("slot_count", 1)} for p in positions]
    else:
        out["donor"] = body.donor
        out["xmatch"] = body.xmatch
    return out


@ws_router.post("/api/shifts/from-template")
def api_create_shift_from_template(template: int = Query(..., ge=1, le=5), name: str | None = Query(None)):
    sid = create_shift_from_template(template, name_override=name)
    shifts = list_shifts()
    created = next((s for s in shifts if s["id"] == sid), None)
    return {"id": sid, "shift": created}


@ws_router.post("/api/apply-template")
def api_apply_template(template: int = Query(..., ge=1, le=6)):
    """Apply template: creates shift(s) and staff (for templates that seed staff)."""
    apply_template(template)
    staff = list_staff()
    all_shifts = list_shifts()
    return {
        "shift_ids": [s["id"] for s in all_shifts],
        "staff_loaded": len(staff) > 0,
        "staff_count": len(staff),
        "shift_count": len(all_shifts),
    }


@ws_router.post("/api/clear-all")
def api_clear_all():
    """ล้างทั้งหมด: บุคลากร, กะ, ตาราง — กลับเป็นหน้าว่าง"""
    clear_all()
    return {"ok": True}


@ws_router.put("/api/shifts/{shift_id:int}")
def api_update_shift(shift_id: int, body: ShiftUpdate):
    positions = None
    if body.positions is not None:
        positions = [{"name": p.name, "constraint_note": p.constraint_note or "", "regular_only": p.regular_only or False, "slot_count": max(1, p.slot_count or 1), "time_window_name": (p.time_window_name or "").strip() or None, "required_skill": (p.required_skill or "").strip() or None, "min_skill_level": max(0, int(p.min_skill_level or 0)), "allowed_titles": list(p.allowed_titles or []), "max_per_week": max(0, int(p.max_per_week or 0)), "active_weekdays": (p.active_weekdays or "").strip() or None, "holiday_mode": (p.holiday_mode or "all").strip() or "all"} for p in body.positions]
    try:
        update_shift(shift_id, body.name, body.donor, body.xmatch, positions=positions, active_days=body.active_days, active_days_of_month=body.active_days_of_month, include_holidays=body.include_holidays, title_requirements=body.title_requirements)
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="ชื่อกะซ้ำ มีอยู่แล้ว กรุณาใช้ชื่ออื่น")
    out = {"id": shift_id, "name": body.name}
    if positions is not None:
        out["positions"] = [{"name": p["name"], "constraint_note": p["constraint_note"], "regular_only": p["regular_only"], "slot_count": p.get("slot_count", 1)} for p in positions]
    else:
        out["donor"] = body.donor
        out["xmatch"] = body.xmatch
    return out


@ws_router.delete("/api/shifts/{shift_id:int}")
def api_delete_shift(shift_id: int):
    delete_shift(shift_id)
    return {"ok": True}


@ws_router.post("/api/shifts/{shift_id:int}/move")
def api_move_shift(shift_id: int, body: ShiftMoveBody):
    direction = (body.direction or "").strip().lower()
    if direction not in ("up", "down"):
        raise HTTPException(status_code=400, detail="direction ต้องเป็น 'up' หรือ 'down'")
    try:
        moved = move_shift(shift_id, direction)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return {"ok": True, "moved": moved}


# --- API: Settings ---
@ws_router.get("/api/settings")
def api_get_settings():
    return {
        "num_days": get_num_days(),
        "schedule_start_date": get_schedule_start_date() or "",
        "holiday_dates": get_holiday_dates(),
    }


@ws_router.get("/api/settings/num_days")
def api_get_num_days():
    return {"value": get_num_days()}


@ws_router.put("/api/settings/num_days")
def api_set_num_days(body: NumDaysUpdate):
    set_num_days(body.value)
    return {"value": body.value}


@ws_router.put("/api/settings")
def api_set_settings(body: SettingsUpdate):
    if body.num_days is not None:
        set_num_days(body.num_days)
    if body.schedule_start_date is not None:
        set_schedule_start_date(body.schedule_start_date.strip() or "")
    if body.holiday_dates is not None:
        set_holiday_dates(body.holiday_dates.strip())
    return api_get_settings()


# --- API: Schedule (run + get) ---
@ws_router.post("/api/schedule/run")
def api_run_schedule(
    body: ScheduleRunBody | None = Body(None),
    num_days_q: int | None = Query(None, alias="num_days"),
    schedule_start_date_q: str | None = Query(None, alias="schedule_start_date"),
):
    num_days_val = (body.num_days if body is not None and body.num_days is not None else None) or num_days_q
    start_val = (body.schedule_start_date.strip() if body and body.schedule_start_date and body.schedule_start_date.strip() else None) or (schedule_start_date_q.strip() if schedule_start_date_q and schedule_start_date_q.strip() else None)
    if num_days_val is not None:
        set_num_days(num_days_val)
    if start_val is not None:
        set_schedule_start_date(start_val)
    num_days = num_days_val or get_num_days()
    mt_list = get_mt_list()
    shift_list = get_shift_list()
    if not mt_list:
        raise HTTPException(status_code=400, detail="No staff. Add at least one staff.")
    if not shift_list:
        raise HTTPException(status_code=400, detail="No shifts. Add at least one shift.")
    start_date_str = get_schedule_start_date()
    logger.info("Running schedule: num_days=%d start_date=%s staff=%d shifts=%d",
                num_days, start_date_str, len(mt_list), len(shift_list))
    try:
        slots, solver, status = generate_schedule(num_days=num_days, start_date_str=start_date_str or None)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # กรณี solver fail จริงๆ (MODEL_INVALID ฯลฯ) — ไม่ใช่แค่ infeasible
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        reasons = diagnose_infeasible(mt_list, shift_list, num_days, start_date_str)
        logger.warning("Schedule infeasible: %s", reasons)
        raise HTTPException(
            status_code=422,
            detail={"message": "Solver could not find any solution.", "reasons": reasons},
        )

    start_date = get_schedule_start_date()
    run_id = save_schedule(num_days, slots, start_date=start_date)
    increment_schedule_run_count()
    data = get_schedule(run_id)

    # ถ้ามี dummy slots → แจ้งเตือน + วิเคราะห์สาเหตุ
    dummy_slots = [s for s in slots if s.get("is_dummy")]
    real_slots = [s for s in slots if not s.get("is_dummy")]
    result = {"run_id": run_id, "schedule": data}
    if dummy_slots:
        hints = diagnose_infeasible(mt_list, shift_list, num_days, start_date_str)
        logger.warning("Schedule has %d dummy slots. Hints: %s", len(dummy_slots), hints)
        result["has_dummy"] = True
        result["dummy_count"] = len(dummy_slots)
        result["infeasibility_hints"] = hints
    else:
        result["has_dummy"] = False
        result["dummy_count"] = 0

    # นับคนที่ถูกจัดมากกว่า 1 เวร/วัน
    from collections import Counter
    day_counts = Counter((s["staff_name"], s["day"]) for s in real_slots)
    multi_shift_cases = {k: v for k, v in day_counts.items() if v > 1}
    result["multi_shift_count"] = len(multi_shift_cases)
    if multi_shift_cases:
        # สรุปให้ frontend แสดง
        details = []
        for (name, day), count in sorted(multi_shift_cases.items(), key=lambda x: (x[0][1], x[0][0])):
            details.append({"staff_name": name, "day": day, "shifts_on_day": count})
        result["multi_shift_details"] = details

    logger.info("Schedule run_id=%d saved (slots=%d dummy=%d multi_shift=%d)", run_id, len(slots), len(dummy_slots), len(multi_shift_cases))
    return result


@ws_router.get("/api/schedule/run/stream")
def api_run_schedule_stream(
    num_days_q: int | None = Query(None, alias="num_days"),
    schedule_start_date_q: str | None = Query(None, alias="schedule_start_date"),
):
    """SSE endpoint: streams solver progress then final result."""
    import json as _json, queue as _queue, threading as _threading

    num_days_val = num_days_q
    start_val = schedule_start_date_q.strip() if schedule_start_date_q and schedule_start_date_q.strip() else None
    if num_days_val is not None:
        set_num_days(num_days_val)
    if start_val is not None:
        set_schedule_start_date(start_val)
    num_days = num_days_val or get_num_days()
    mt_list_check = get_mt_list()
    shift_list_check = get_shift_list()
    if not mt_list_check:
        raise HTTPException(status_code=400, detail="No staff.")
    if not shift_list_check:
        raise HTTPException(status_code=400, detail="No shifts.")
    start_date_str = get_schedule_start_date()

    progress_q: _queue.Queue = _queue.Queue()

    # Capture workspace context for the solver thread
    from database import _workspace_db_path
    _ws_path = _workspace_db_path.get(None)

    def _on_progress(info):
        progress_q.put(("progress", info))

    def _solver_thread():
        if _ws_path is not None:
            _workspace_db_path.set(_ws_path)
        try:
            slots, solver_obj, status = generate_schedule(
                num_days=num_days, start_date_str=start_date_str or None, on_progress=_on_progress,
            )
            progress_q.put(("done", (slots, status)))
        except Exception as exc:
            progress_q.put(("error", str(exc)))

    _threading.Thread(target=_solver_thread, daemon=True).start()

    def _event_stream():
        from ortools.sat.python import cp_model as _cp
        while True:
            try:
                kind, data = progress_q.get(timeout=1)
            except _queue.Empty:
                yield "event: ping\ndata: {}\n\n"
                continue
            if kind == "progress":
                yield f"event: progress\ndata: {_json.dumps(data)}\n\n"
            elif kind == "error":
                yield f"event: error\ndata: {_json.dumps({'message': data})}\n\n"
                return
            elif kind == "done":
                slots, status = data
                if status not in (_cp.OPTIMAL, _cp.FEASIBLE):
                    reasons = diagnose_infeasible(get_mt_list(), get_shift_list(), num_days, start_date_str)
                    yield f"event: error\ndata: {_json.dumps({'message': 'Solver could not find any solution.', 'reasons': reasons})}\n\n"
                    return
                start_date = get_schedule_start_date()
                run_id = save_schedule(num_days, slots, start_date=start_date)
                increment_schedule_run_count()
                sched_data = get_schedule(run_id)
                dummy_slots = [s for s in slots if s.get("is_dummy")]
                real_slots = [s for s in slots if not s.get("is_dummy")]
                result = {"run_id": run_id, "schedule": sched_data, "has_dummy": bool(dummy_slots), "dummy_count": len(dummy_slots)}
                if dummy_slots:
                    result["infeasibility_hints"] = diagnose_infeasible(get_mt_list(), get_shift_list(), num_days, start_date_str)
                day_counts = collections.Counter((s["staff_name"], s["day"]) for s in real_slots)
                multi = {k: v for k, v in day_counts.items() if v > 1}
                result["multi_shift_count"] = len(multi)
                if multi:
                    result["multi_shift_details"] = [
                        {"staff_name": n, "day": d, "shifts_on_day": c}
                        for (n, d), c in sorted(multi.items(), key=lambda x: (x[0][1], x[0][0]))
                    ]
                yield f"event: progress\ndata: {_json.dumps({'percent': 100, 'solutions': 0})}\n\n"
                yield f"event: result\ndata: {_json.dumps(result)}\n\n"
                return

    return StreamingResponse(_event_stream(), media_type="text/event-stream")


@ws_router.get("/api/schedule/latest")
def api_get_latest_schedule():
    data = get_latest_schedule()
    if data is None:
        raise HTTPException(status_code=404, detail="No schedule yet. Run the scheduler first.")
    return data


@ws_router.get("/api/schedule/{run_id:int}")
def api_get_schedule(run_id: int):
    data = get_schedule(run_id)
    if data is None:
        raise HTTPException(status_code=404, detail="Schedule not found.")
    return data


# --- Export CSV ---
@ws_router.get("/api/schedule/export/csv")
def api_export_schedule_csv(run_id: int | None = None):
    from datetime import datetime, timedelta
    if run_id is not None:
        data = get_schedule(run_id)
    else:
        data = get_latest_schedule()
    if data is None:
        raise HTTPException(status_code=404, detail="No schedule to export.")

    slots = data.get("slots", [])
    start_date = data.get("start_date")
    pos_key = "position" if slots and "position" in slots[0] else "room"

    # Build ordered shift/position/slot columns from slots
    seen_cols: dict = {}  # (shift_name, pos, slot_index) -> col_header
    for s in slots:
        pos = s.get(pos_key) or s.get("room") or ""
        si = s.get("slot_index", 0)
        key = (s["shift_name"], pos, si)
        if key not in seen_cols:
            label = f"{s['shift_name']} / {pos}" if pos else s["shift_name"]
            if si > 0:
                label += f" ({si + 1})"
            seen_cols[key] = label

    col_keys = list(seen_cols.keys())
    col_headers = [seen_cols[k] for k in col_keys]

    # Index slots by (day, shift, pos, slot_index)
    slot_map: dict = {}
    for s in slots:
        pos = s.get(pos_key) or s.get("room") or ""
        si = s.get("slot_index", 0)
        slot_map[(s["day"], s["shift_name"], pos, si)] = s

    num_days = data.get("num_days") or (max(s["day"] for s in slots) + 1 if slots else 0)

    # Resolve base date
    base = None
    if start_date:
        try:
            base = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            pass

    THAI_MONTHS_SHORT = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
                         "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
    DAY_TH = ["จ", "อ", "พ", "พฤ", "ศ", "ส", "อา"]

    buf = io.StringIO()
    w = csv.writer(buf)

    # Header row
    w.writerow(["วันที่", "วัน"] + col_headers)

    for day in range(num_days):
        if base:
            dt = base + timedelta(days=day)
            date_label = f"{dt.day} {THAI_MONTHS_SHORT[dt.month - 1]} {dt.year + 543}"
            day_label = DAY_TH[dt.weekday()]
        else:
            date_label = str(day + 1)
            day_label = ""

        row = [date_label, day_label]
        for (sn, pos, si) in col_keys:
            s = slot_map.get((day, sn, pos, si))
            if s is None:
                row.append("—")  # shift inactive this day
            elif s.get("is_dummy"):
                row.append("ว่าง")
            else:
                row.append(s.get("staff_name", ""))
        w.writerow(row)

    # Blank separator + summary section
    w.writerow([])
    w.writerow(["สรุปเวรต่อคน"])
    w.writerow(["ชื่อ", "จำนวนเวร"])
    count_by_staff: dict = {}
    for s in slots:
        if not s.get("is_dummy"):
            count_by_staff[s["staff_name"]] = count_by_staff.get(s["staff_name"], 0) + 1
    for name, cnt in sorted(count_by_staff.items(), key=lambda x: -x[1]):
        w.writerow([name, cnt])

    content = buf.getvalue().encode("utf-8-sig")
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=schedule.csv"},
    )


@ws_router.get("/api/schedule/export/xlsx")
def api_export_schedule_xlsx(run_id: int | None = None):
    from datetime import datetime, timedelta
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter

    if run_id is not None:
        data = get_schedule(run_id)
    else:
        data = get_latest_schedule()
    if data is None:
        raise HTTPException(status_code=404, detail="No schedule to export.")

    slots = data.get("slots", [])
    start_date = data.get("start_date")
    pos_key = "position" if slots and "position" in slots[0] else "room"

    THAI_MONTHS_SHORT = ["ม.ค.","ก.พ.","มี.ค.","เม.ย.","พ.ค.","มิ.ย.",
                         "ก.ค.","ส.ค.","ก.ย.","ต.ค.","พ.ย.","ธ.ค."]
    DAY_TH = ["จ","อ","พ","พฤ","ศ","ส","อา"]
    WEEKEND = {5, 6}  # Saturday, Sunday

    # Shift color palette (cool tones, matches print)
    SHIFT_BG = ["EFF6FF","F0FDFA","EEF2FF","F0F9FF","FAF5FF","ECFEFF","F0FDF4","F5F3FF"]

    base = None
    if start_date:
        try:
            base = datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            pass

    # Build ordered columns
    seen_cols: dict = {}
    shift_order: list = []  # preserve original shift order
    for s in slots:
        pos = s.get(pos_key) or s.get("room") or ""
        si = s.get("slot_index", 0)
        key = (s["shift_name"], pos, si)
        if key not in seen_cols:
            label = f"{s['shift_name']} / {pos}" if pos else s["shift_name"]
            if si > 0:
                label += f" ({si+1})"
            seen_cols[key] = label
        if s["shift_name"] not in shift_order:
            shift_order.append(s["shift_name"])
    # Sort col_keys so all positions of the same shift are consecutive
    col_keys = sorted(seen_cols.keys(), key=lambda k: (shift_order.index(k[0]), k[1], k[2]))
    col_headers = [seen_cols[k] for k in col_keys]

    # Map shift name → color index
    shift_color_idx: dict = {}
    ci = 0
    for (sn, _, __) in col_keys:
        if sn not in shift_color_idx:
            shift_color_idx[sn] = ci % len(SHIFT_BG)
            ci += 1

    slot_map: dict = {}
    for s in slots:
        pos = s.get(pos_key) or s.get("room") or ""
        si = s.get("slot_index", 0)
        slot_map[(s["day"], s["shift_name"], pos, si)] = s

    num_days = data.get("num_days") or (max(s["day"] for s in slots) + 1 if slots else 0)

    # --- Helpers ---
    def thin_border():
        t = Side(style="thin", color="CBD5E1")
        return Border(left=t, right=t, top=t, bottom=t)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    wb = Workbook()

    # ═══════════════════════════════════════
    # Sheet 1 — ตารางเวร
    # ═══════════════════════════════════════
    ws = wb.active
    ws.title = "ตารางเวร"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(col_keys))
    title_cell = ws.cell(1, 1)
    date_range = ""
    if base and num_days:
        end_dt = base + timedelta(days=num_days - 1)
        date_range = (f"{base.day} {THAI_MONTHS_SHORT[base.month-1]} {base.year+543}"
                      f" – {end_dt.day} {THAI_MONTHS_SHORT[end_dt.month-1]} {end_dt.year+543}")
    title_cell.value = f"ตารางเวร{' · ' + date_range if date_range else ''}"
    title_cell.font = Font(name="Cordia New", bold=True, size=14, color="1D4ED8")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.fill = fill("FFFFFF")
    ws.row_dimensions[1].height = 22

    # ── Headers: 2 rows (row 2 = shift group, row 3 = position) ──
    HDR_FILL = fill("1E293B")
    HDR_FONT = Font(name="Cordia New", bold=True, color="F8FAFC", size=11)
    POS_FONT = Font(name="Cordia New", bold=True, color="1E293B", size=10)
    CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BDR = thin_border()

    def hdr(row, col, val, fnt, flll, aln=None):
        c = ws.cell(row, col, val)
        c.font = fnt
        c.fill = flll
        c.alignment = aln or CTR
        c.border = BDR
        return c

    # วันที่ / วัน — span both header rows
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    hdr(2, 1, "วันที่", HDR_FONT, HDR_FILL)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    hdr(2, 2, "วัน", HDR_FONT, HDR_FILL)

    # Group col_keys by shift name to build merged shift header
    shift_groups: list = []  # (shift_name, start_col, end_col)
    cur_shift = None
    cur_start = 3
    for ci_off, (sn, pos, si) in enumerate(col_keys):
        col_ci = ci_off + 3
        if sn != cur_shift:
            if cur_shift is not None:
                shift_groups.append((cur_shift, cur_start, col_ci - 1))
            cur_shift = sn
            cur_start = col_ci
    if cur_shift is not None:
        shift_groups.append((cur_shift, cur_start, len(col_keys) + 2))

    # Row 2 — shift name (merged, dark)
    for sn, sc, ec in shift_groups:
        if sc == ec:
            ws.cell(2, sc)  # single col, no merge needed
        else:
            ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
        hdr(2, sc, sn, HDR_FONT, HDR_FILL)

    # Row 3 — position name (light background, per shift color)
    for ci_off, (sn, pos, si) in enumerate(col_keys):
        col_ci = ci_off + 3
        label = pos if pos else sn
        if si > 0:
            label += f" {si+1}"
        bg = SHIFT_BG[shift_color_idx[sn]]
        pos_fill = fill(bg)
        pos_border_top = Side(style="medium", color="1E293B")
        c = ws.cell(3, col_ci, label)
        c.font = POS_FONT
        c.fill = pos_fill
        c.alignment = CTR
        from openpyxl.styles import Border as OBorder
        t = Side(style="thin", color="CBD5E1")
        c.border = OBorder(left=t, right=t, top=pos_border_top, bottom=t)

    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 5
    for ci in range(len(col_keys)):
        ws.column_dimensions[get_column_letter(ci + 3)].width = 14

    ws.freeze_panes = "A4"

    # Data rows
    for day in range(num_days):
        row_num = day + 4
        if base:
            dt = base + timedelta(days=day)
            date_label = f"{dt.day} {THAI_MONTHS_SHORT[dt.month-1]} {dt.year+543}"
            day_label = DAY_TH[dt.weekday()]
            is_weekend = dt.weekday() in WEEKEND
        else:
            date_label = str(day + 1)
            day_label = ""
            is_weekend = False

        # Check holiday
        iso = (base + timedelta(days=day)).isoformat() if base else ""
        is_holiday = iso in (data.get("holidays") or [])
        row_base_fill = fill("FEF08A") if is_holiday else (fill("FFF7ED") if is_weekend else None)

        def set_day_cell(col, val, bold=False):
            c = ws.cell(row_num, col, val)
            c.font = Font(name="Cordia New", bold=bold, size=10)
            c.alignment = CTR
            c.border = BDR
            if row_base_fill:
                c.fill = row_base_fill
            elif day % 2 == 1:
                c.fill = fill("F8FAFC")
            return c

        set_day_cell(1, date_label, bold=is_holiday or is_weekend)
        set_day_cell(2, day_label, bold=True)

        for ci, (sn, pos, si) in enumerate(col_keys, start=3):
            s = slot_map.get((day, sn, pos, si))
            c = ws.cell(row_num, ci)
            c.alignment = CTR
            c.border = BDR
            c.font = Font(name="Cordia New", size=10)

            if s is None:
                c.value = ""
                c.fill = fill("F8FAFC")
            elif s.get("is_dummy"):
                c.value = "ว่าง"
                c.fill = fill("FEE2E2")
                c.font = Font(name="Cordia New", size=10, bold=True, color="B91C1C")
            else:
                c.value = s.get("staff_name", "")
                if row_base_fill:
                    c.fill = row_base_fill
                elif day % 2 == 1:
                    c.fill = fill("F8FAFC")
                else:
                    c.fill = fill(SHIFT_BG[shift_color_idx[sn]])

        ws.row_dimensions[row_num].height = 16

    # ═══════════════════════════════════════
    # Sheet 2 — สรุปเวรต่อคน
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("สรุปเวรต่อคน")

    count_by_staff: dict = {}
    shift_matrix: dict = {}
    all_shifts = list(dict.fromkeys(s["shift_name"] for s in slots))
    for s in slots:
        if not s.get("is_dummy"):
            nm = s["staff_name"]
            count_by_staff[nm] = count_by_staff.get(nm, 0) + 1
            k = (nm, s["shift_name"])
            shift_matrix[k] = shift_matrix.get(k, 0) + 1

    sorted_staff = sorted(count_by_staff.items(), key=lambda x: -x[1])
    max_c = sorted_staff[0][1] if sorted_staff else 1

    # Title
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(all_shifts))
    t2 = ws2.cell(1, 1, f"สรุปเวรต่อคน{' · ' + date_range if date_range else ''}")
    t2.font = Font(name="Cordia New", bold=True, size=14, color="1D4ED8")
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 22

    # Header
    for ci, val in enumerate(["#", "ชื่อ", "เวรรวม"] + all_shifts, start=1):
        c = ws2.cell(2, ci, val)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CTR
        c.border = BDR
    ws2.row_dimensions[2].height = 22
    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 10
    for ci in range(len(all_shifts)):
        ws2.column_dimensions[get_column_letter(ci + 4)].width = 12
    ws2.freeze_panes = "A3"

    for i, (name, total) in enumerate(sorted_staff, start=1):
        r = i + 2
        row_fill = fill("F8FAFC") if i % 2 == 0 else None
        for ci, val in enumerate([i, name, total] + [shift_matrix.get((name, sn), 0) or "—" for sn in all_shifts], start=1):
            c = ws2.cell(r, ci, val)
            c.font = Font(name="Cordia New", size=10, bold=(ci == 2))
            c.alignment = Alignment(horizontal="center" if ci != 2 else "left", vertical="center")
            c.border = BDR
            if row_fill:
                c.fill = row_fill
        ws2.row_dimensions[r].height = 16

    # Footer total row
    fr = len(sorted_staff) + 3
    for ci, val in enumerate(["", "รวม", sum(count_by_staff.values())] + [
        sum(shift_matrix.get((n, sn), 0) for n, _ in sorted_staff) for sn in all_shifts
    ], start=1):
        c = ws2.cell(fr, ci, val)
        c.font = Font(name="Cordia New", bold=True, size=10)
        c.alignment = CTR
        c.border = BDR
        c.fill = fill("E2E8F0")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=schedule.xlsx"},
    )


# --- Manual Slot Assignment ---
@ws_router.patch("/api/schedule/{run_id:int}/slot")
def api_assign_slot(run_id: int, body: SlotAssign):
    """Manual override: กำหนด staff ให้ slot ที่ระบุ (ใช้แทน dummy หรือสลับคน)"""
    data = get_schedule(run_id)
    if data is None:
        raise HTTPException(status_code=404, detail="Schedule not found.")
    name = body.staff_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="staff_name ต้องไม่ว่าง")
    warnings = _check_staff_off_day_warnings(name, body.day)
    try:
        update_slot_staff(run_id, body.day, body.shift_name, body.position, body.slot_index, name)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    updated = get_schedule(run_id)
    return {"ok": True, "run_id": run_id, "schedule": updated, "warnings": warnings}


# --- Manual Slot Swap ---
@ws_router.post("/api/schedule/{run_id:int}/swap")
def api_swap_slots(run_id: int, body: SlotSwap):
    """สลับ staff ระหว่างสอง slot (atomic, ไม่ block โดย same-day guard)"""
    data = get_schedule(run_id)
    if data is None:
        raise HTTPException(status_code=404, detail="Schedule not found.")
    # ดึงชื่อคนก่อน swap เพื่อเช็ค off-day warnings
    data_before = get_schedule(run_id)
    warnings = []
    if data_before:
        slots = data_before if isinstance(data_before, list) else data_before.get("slots", [])
        for s in slots:
            if (s.get("day") == body.day_a and s.get("shift_name") == body.shift_name_a
                    and s.get("position") == body.position_a and s.get("slot_index") == body.slot_index_a):
                # คนนี้จะไป day_b
                warnings.extend(_check_staff_off_day_warnings(s["staff_name"], body.day_b))
            if (s.get("day") == body.day_b and s.get("shift_name") == body.shift_name_b
                    and s.get("position") == body.position_b and s.get("slot_index") == body.slot_index_b):
                # คนนี้จะไป day_a
                warnings.extend(_check_staff_off_day_warnings(s["staff_name"], body.day_a))
    try:
        swap_slots(
            run_id,
            body.day_a, body.shift_name_a, body.position_a, body.slot_index_a,
            body.day_b, body.shift_name_b, body.position_b, body.slot_index_b,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    updated = get_schedule(run_id)
    return {"ok": True, "run_id": run_id, "schedule": updated, "warnings": warnings}


# --- API: Staff Pairs ---
@ws_router.get("/api/staff-pairs")
def api_list_staff_pairs():
    return list_staff_pairs()


@ws_router.post("/api/staff-pairs")
def api_add_staff_pair(body: StaffPairCreate):
    if body.staff_id_1 == body.staff_id_2:
        raise HTTPException(status_code=400, detail="ต้องเลือกคนละคน")
    if body.pair_type not in ("together", "apart", "depends_on"):
        raise HTTPException(status_code=400, detail="pair_type ต้องเป็น 'together', 'apart' หรือ 'depends_on'")
    add_staff_pair(body.staff_id_1, body.staff_id_2, body.pair_type, shift_names=body.shift_names)
    return {"ok": True}


@ws_router.post("/api/staff-pairs/batch")
def api_add_staff_pairs_batch(body: list[StaffPairBatchItem]):
    """เพิ่มหลายคู่พร้อมกัน รองรับ name_1, name_2 (ชื่อบุคลากร)"""
    from database import list_staff

    staff_by_name = {s["name"]: s["id"] for s in list_staff()}
    errors = []
    for i, item in enumerate(body):
        if not item.name_1 or not item.name_2:
            errors.append(f"แถว {i + 1}: ต้องระบุชื่อทั้ง 2 คน")
            continue
        if item.name_1.strip() == item.name_2.strip():
            errors.append(f"แถว {i + 1}: ต้องเลือกคนละคน")
            continue
        if item.pair_type not in ("together", "apart", "depends_on"):
            errors.append(f"แถว {i + 1}: pair_type ต้องเป็น together, apart หรือ depends_on")
            continue
        if not staff_by_name.get(item.name_1.strip()):
            errors.append(f"แถว {i + 1}: ไม่พบบุคลากร '{item.name_1}'")
            continue
        if not staff_by_name.get(item.name_2.strip()):
            errors.append(f"แถว {i + 1}: ไม่พบบุคลากร '{item.name_2}'")
            continue
    if errors:
        raise HTTPException(status_code=400, detail="; ".join(errors[:10]) + (" ..." if len(errors) > 10 else ""))
    added = 0
    for item in body:
        s1 = staff_by_name[item.name_1.strip()]
        s2 = staff_by_name[item.name_2.strip()]
        add_staff_pair(s1, s2, item.pair_type, shift_names=item.shift_names)
        added += 1
    return {"ok": True, "added": added}


@ws_router.delete("/api/staff-pairs/{pair_id:int}")
def api_remove_staff_pair(pair_id: int):
    remove_staff_pair(pair_id)
    return {"ok": True}


# --- API: Import/Export ---
@ws_router.get("/api/export")
def api_export():
    return export_all_data()


@ws_router.post("/api/import")
def api_import(data: dict = Body(...)):
    import_all_data(data)
    return {"ok": True}


# --- Frontend: serve index.html inside workspace ---
@ws_router.get("/staff", response_class=HTMLResponse)
def staff_page():
    staff_html = STATIC_DIR / "staff.html"
    if staff_html.exists():
        return FileResponse(staff_html)
    return HTMLResponse("<h1>Staff</h1><p>Add static/staff.html for the staff detail page.</p>")


@ws_router.get("/", response_class=HTMLResponse)
def workspace_index():
    index_html = STATIC_DIR / "index.html"
    if index_html.exists():
        return FileResponse(index_html)
    return HTMLResponse("<h1>MT Shift Optimizer</h1><p>Add static/index.html for the UI.</p>")


# Mount workspace router
app.include_router(ws_router, prefix="/w/{workspace_id}")


# ==========================================
# Global routes (not workspace-scoped)
# ==========================================

# --- Workspace management API ---
@app.post("/api/workspaces", dependencies=[Depends(verify_api_key)])
def api_create_workspace(body: WorkspaceCreate = Body(WorkspaceCreate())):
    wid, token = create_workspace(body.name)
    logger.info("Created workspace id=%s name=%r", wid, body.name)
    return {"id": wid, "token": token, "url": f"/w/{wid}/"}


@app.get("/api/workspaces", dependencies=[Depends(verify_api_key)])
def api_list_workspaces():
    # dev mode (no API_KEY) → ส่ง token กลับมาด้วยเพื่อให้ browser sync
    return list_workspaces(include_tokens=_DISABLE_WORKSPACE_AUTH)


@app.get("/api/workspaces/{workspace_id}")
def api_get_workspace(workspace_id: str):
    ws = get_workspace(workspace_id)
    if not ws:
        raise HTTPException(status_code=404, detail="Workspace not found")
    return ws


@app.delete("/api/workspaces/{workspace_id}", dependencies=[Depends(verify_api_key)])
def api_delete_workspace(workspace_id: str):
    ok = delete_workspace(workspace_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Workspace not found")
    logger.info("Deleted workspace id=%s", workspace_id)
    return {"ok": True}


# --- สถิติรวม: จำนวนครั้งที่กดสร้างตารางเวร ---
@app.get("/api/stats/schedule-runs")
def api_get_schedule_run_count():
    """คืนจำนวนครั้งที่กดสร้างตารางเวรรวมทั้งหมด (ทุก workspace)"""
    return {"total": get_schedule_run_count()}


# --- Landing page ---
@app.get("/", response_class=HTMLResponse)
def landing():
    landing_html = STATIC_DIR / "landing.html"
    if landing_html.exists():
        return FileResponse(landing_html)
    # Fallback: redirect to first workspace if exists, or show basic HTML
    workspaces = list_workspaces()
    if workspaces:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=f"/w/{workspaces[0]['id']}/")
    return HTMLResponse("<h1>MT Shift Optimizer</h1><p>No workspaces yet.</p>")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
